# -*- coding: utf-8 -*-
"""
入库/光伏 数据映射器（带结束日期/时间自动判定）
1) 读取河流-水库配置 JSON（包含 rivers / data_type）
2) 为每个水库设置：xlsx/Sheet/列/起始行/起始日期，自动算 end_date（逐日）
3) 可选：添加多个光伏序列（逐小时）：xlsx/Sheet/列/起始行/起始日期时间，自动算 end_datetime
4) 保存为 inflow+pv 映射 JSON

pip install pyqt5 openpyxl
"""

import json
import sys
import os
from typing import List, Dict, Any, Optional
from datetime import date, datetime, timedelta

from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QFormLayout, QLabel, QLineEdit,
    QPushButton, QFileDialog, QComboBox, QSpinBox, QDateEdit, QDateTimeEdit, QMessageBox,
    QScrollArea, QCheckBox, QGroupBox
)
from PyQt5.QtCore import Qt, QDate, QDateTime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtCore import QTime


# ------------------------------ 后端：配置读取与保存 ------------------------------
def load_river_config(cfg_path: str) -> Dict[str, Any]:
    with open(cfg_path, "r", encoding="utf-8") as f:
        return json.load(f)


def extract_reservoir_list(river_cfg: Dict[str, Any]) -> List[Dict[str, str]]:
    """从河流配置中提取所有水库清单。"""
    reservoirs = []
    for r in river_cfg.get("rivers", []):
        river_name = r.get("river_name", "")
        for res in r.get("reservoirs", []):
            reservoirs.append({"river_name": river_name, "reservoir_name": str(res)})
    return reservoirs


def save_mapping(save_path: str,
                 data_type: str,
                 inflow_mappings: List[Dict[str, Any]],
                 pv_mappings: List[Dict[str, Any]]):
    """
    保存最终 JSON 映射。
    - data_type 为项目级别（水库入库尺度），通常是 daily/hourly；光伏固定逐小时。
    """
    out = {
        "data_type": data_type,
        "reservoir_inflow_sources": inflow_mappings,  # 逐日或逐小时，但通常逐日
        "pv_sources": pv_mappings                     # 逐小时
    }
    with open(save_path, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=4, ensure_ascii=False)


# ------------------------------ 公共：Excel 扫描工具 ------------------------------
def find_last_non_empty_row(xlsx_path: str, sheet_name: str, col_index: int) -> Optional[int]:
    """
    返回该列中最后一个非空单元格的行号；若无法确定返回 None。
    允许中间有空值，但以自下而上找到的第一个非空为准。
    """
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        # 自底向上找第一个非空
        for r in range(max_row, 0, -1):
            v = ws.cell(row=r, column=col_index).value
            if v is not None and str(v).strip() != "":
                return r
        return None
    except Exception:
        return None


# ------------------------------ 前端：水库行（逐日） ------------------------------
class ReservoirRow(QWidget):
    def __init__(self, river_name: str, reservoir_name: str, parent=None):
        super().__init__(parent)
        self.river_name = river_name
        self.reservoir_name = reservoir_name

        # 左侧标签
        info_col = QFormLayout()
        self.lbl_river = QLabel(river_name)
        self.lbl_res = QLabel(reservoir_name)
        info_col.addRow("河流：", self.lbl_river)
        info_col.addRow("水库：", self.lbl_res)

        # 文件与表
        self.file_edit = QLineEdit()
        self.file_edit.setPlaceholderText("请选择 .xlsx 文件路径")
        self.file_edit.setReadOnly(True)
        self.btn_browse = QPushButton("浏览...")
        self.btn_browse.clicked.connect(self.on_browse_file)

        file_row = QHBoxLayout()
        file_row.addWidget(self.file_edit, 1)
        file_row.addWidget(self.btn_browse)

        self.sheet_combo = QComboBox()
        self.sheet_combo.currentIndexChanged.connect(self.on_sheet_changed)

        self.col_combo = QComboBox()  # 存 userData={"letter":, "index":}
        self.col_combo.currentIndexChanged.connect(self._auto_calc_end)

        self.row_spin = QSpinBox()
        self.row_spin.setRange(1, 10**7)
        self.row_spin.setValue(1)
        self.row_spin.valueChanged.connect(self._auto_calc_end)

        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setDate(QDate(2000, 1, 1))
        self.date_edit.dateChanged.connect(self._auto_calc_end)

        # 自动结束日期显示
        self.end_label = QLabel("—")
        self.end_label.setToolTip("自动根据列数据长度与起始日期推算")

        center_form = QFormLayout()
        center_form.addRow("Excel 文件：", QWidget())  # 占位
        center_form.addRow(file_row)
        center_form.addRow("Sheet：", self.sheet_combo)
        center_form.addRow("列：", self.col_combo)
        center_form.addRow("起始行：", self.row_spin)
        center_form.addRow("起始日期：", self.date_edit)
        center_form.addRow("结束日期（自动）：", self.end_label)

        # 总体布局
        root = QHBoxLayout(self)
        root.addLayout(info_col, 0)
        root.addSpacing(12)
        root.addLayout(center_form, 1)

        self._last_wb_path = None

    # ---------- 交互 ----------
    def on_browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择入库径流 Excel 文件（.xlsx）", "", "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        self._last_wb_path = path
        self.file_edit.setText(path)
        self._load_sheet_names(path)
        self._refresh_columns()
        self._auto_calc_end()

    def on_sheet_changed(self, _idx: int):
        self._refresh_columns()
        self._auto_calc_end()

    # ---------- 辅助 ----------
    def _load_sheet_names(self, xlsx_path: str):
        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            self.sheet_combo.clear()
            if not sheets:
                self.sheet_combo.addItem("(空工作簿)")
                self.sheet_combo.setEnabled(False)
            else:
                self.sheet_combo.addItems(sheets)
                self.sheet_combo.setEnabled(True)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取工作簿失败：\n{xlsx_path}\n{e}")
            self.sheet_combo.clear()
            self.col_combo.clear()
            self._last_wb_path = None

    def _refresh_columns(self):
        self.col_combo.clear()
        if not self._last_wb_path:
            return
        sheet = self.sheet_combo.currentText().strip()
        if not sheet:
            return
        try:
            wb = load_workbook(self._last_wb_path, read_only=True, data_only=True)
            if sheet not in wb.sheetnames:
                return
            ws = wb[sheet]
            max_col = ws.max_column or 1
            for i in range(1, max_col + 1):
                letter = get_column_letter(i)
                self.col_combo.addItem(f"{letter}（第{i}列）", userData={"letter": letter, "index": i})
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取列信息失败：\n{self._last_wb_path}\n{e}")

    def _auto_calc_end(self):
        """根据列最后非空行 + 起始日期 计算结束日期（逐日）。"""
        self.end_label.setText("—")
        if not self._last_wb_path:
            return
        sheet = self.sheet_combo.currentText().strip()
        data = self.col_combo.currentData() or {}
        col_index = data.get("index")
        if not sheet or not col_index:
            return
        start_row = int(self.row_spin.value())
        last_row = find_last_non_empty_row(self._last_wb_path, sheet, col_index)
        if last_row is None or last_row < start_row:
            self.end_label.setText("无法判定/无数据")
            return
        n = last_row - start_row + 1
        start_qdate: QDate = self.date_edit.date()
        start_pydate = date(start_qdate.year(), start_qdate.month(), start_qdate.day())
        end_pydate = start_pydate + timedelta(days=max(n - 1, 0))
        self.end_label.setText(end_pydate.isoformat())

    # ---------- 导出 ----------
    def to_mapping(self) -> Dict[str, Any]:
        data = self.col_combo.currentData() or {}
        col_letter = data.get("letter")
        col_index = data.get("index")
        start_date_str = self.date_edit.date().toString("yyyy-MM-dd")
        end_date_str = self.end_label.text().strip() if self.end_label.text().strip() not in ("—", "无法判定/无数据") else None

        # 计算 record_count（基于列最后非空行）
        record_count = None
        if self._last_wb_path and col_index and self.sheet_combo.currentText().strip():
            last_row = find_last_non_empty_row(self._last_wb_path, self.sheet_combo.currentText().strip(), col_index)
            if last_row is not None and last_row >= int(self.row_spin.value()):
                record_count = last_row - int(self.row_spin.value()) + 1

        return {
            "time_scale": "daily",
            "river_name": self.river_name,
            "reservoir_name": self.reservoir_name,
            "file_path": self.file_edit.text().strip(),
            "sheet_name": self.sheet_combo.currentText().strip(),
            "column_letter": col_letter,
            "column_index": col_index,
            "start_row": int(self.row_spin.value()),
            "start_date": start_date_str,
            "end_date": end_date_str,
            "record_count": record_count
        }

    def is_valid(self) -> bool:
        m = self.to_mapping()
        # 若选择了文件，则要求必要项齐备
        if m["file_path"]:
            return all([m["sheet_name"], m["column_letter"], m["column_index"], m["start_date"]])
        return True


# ------------------------------ 前端：光伏行（逐小时） ------------------------------
class PVRow(QWidget):
    def __init__(self, pv_index: int, parent=None):
        super().__init__(parent)
        self.pv_name_edit = QLineEdit(f"PV{pv_index}")
        self.pv_name_edit.setPlaceholderText("光伏序列名（可修改）")

        self.file_edit = QLineEdit()
        self.file_edit.setPlaceholderText("请选择 .xlsx 文件路径")
        self.file_edit.setReadOnly(True)
        self.btn_browse = QPushButton("浏览...")
        self.btn_browse.clicked.connect(self.on_browse_file)

        file_row = QHBoxLayout()
        file_row.addWidget(self.file_edit, 1)
        file_row.addWidget(self.btn_browse)

        self.sheet_combo = QComboBox()
        self.sheet_combo.currentIndexChanged.connect(self.on_sheet_changed)

        self.col_combo = QComboBox()
        self.col_combo.currentIndexChanged.connect(self._auto_calc_end)

        self.row_spin = QSpinBox()
        self.row_spin.setRange(1, 10**7)
        self.row_spin.setValue(1)
        self.row_spin.valueChanged.connect(self._auto_calc_end)

        self.dt_edit = QDateTimeEdit()
        self.dt_edit.setCalendarPopup(True)
        self.dt_edit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self.dt_edit.setDateTime(QDateTime(QDate(2000, 1, 1), QTime(0, 0)))
        self.dt_edit.dateTimeChanged.connect(self._auto_calc_end)

        self.end_label = QLabel("—")
        self.end_label.setToolTip("自动根据列数据长度与起始时间推算（逐小时）")

        form = QFormLayout(self)
        form.addRow("序列名称：", self.pv_name_edit)
        form.addRow("Excel 文件：", QWidget())
        form.addRow(file_row)
        form.addRow("Sheet：", self.sheet_combo)
        form.addRow("列：", self.col_combo)
        form.addRow("起始行：", self.row_spin)
        form.addRow("起始日期时间：", self.dt_edit)
        form.addRow("结束时间（自动）：", self.end_label)

        self._last_wb_path = None

    def on_browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择光伏 Excel 文件（.xlsx）", "", "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        self._last_wb_path = path
        self.file_edit.setText(path)
        self._load_sheet_names(path)
        self._refresh_columns()
        self._auto_calc_end()

    def on_sheet_changed(self, _idx: int):
        self._refresh_columns()
        self._auto_calc_end()

    def _load_sheet_names(self, xlsx_path: str):
        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            self.sheet_combo.clear()
            if not sheets:
                self.sheet_combo.addItem("(空工作簿)")
                self.sheet_combo.setEnabled(False)
            else:
                self.sheet_combo.addItems(sheets)
                self.sheet_combo.setEnabled(True)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取工作簿失败：\n{xlsx_path}\n{e}")
            self.sheet_combo.clear()
            self.col_combo.clear()
            self._last_wb_path = None

    def _refresh_columns(self):
        self.col_combo.clear()
        if not self._last_wb_path:
            return
        sheet = self.sheet_combo.currentText().strip()
        if not sheet:
            return
        try:
            wb = load_workbook(self._last_wb_path, read_only=True, data_only=True)
            if sheet not in wb.sheetnames:
                return
            ws = wb[sheet]
            max_col = ws.max_column or 1
            for i in range(1, max_col + 1):
                letter = get_column_letter(i)
                self.col_combo.addItem(f"{letter}（第{i}列）", userData={"letter": letter, "index": i})
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取列信息失败：\n{self._last_wb_path}\n{e}")

    def _auto_calc_end(self):
        """根据列最后非空行 + 起始时间 计算结束时间（逐小时）。"""
        self.end_label.setText("—")
        if not self._last_wb_path:
            return
        sheet = self.sheet_combo.currentText().strip()
        data = self.col_combo.currentData() or {}
        col_index = data.get("index")
        if not sheet or not col_index:
            return
        start_row = int(self.row_spin.value())
        last_row = find_last_non_empty_row(self._last_wb_path, sheet, col_index)
        if last_row is None or last_row < start_row:
            self.end_label.setText("无法判定/无数据")
            return
        n = last_row - start_row + 1
        start_qdt: QDateTime = self.dt_edit.dateTime()
        start_dt = datetime(
            start_qdt.date().year(), start_qdt.date().month(), start_qdt.date().day(),
            start_qdt.time().hour(), start_qdt.time().minute(), start_qdt.time().second()
        )
        end_dt = start_dt + timedelta(hours=max(n - 1, 0))
        self.end_label.setText(end_dt.strftime("%Y-%m-%d %H:%M:%S"))

    def to_mapping(self) -> Dict[str, Any]:
        data = self.col_combo.currentData() or {}
        col_letter = data.get("letter")
        col_index = data.get("index")
        start_qdt: QDateTime = self.dt_edit.dateTime()
        start_dt = datetime(
            start_qdt.date().year(), start_qdt.date().month(), start_qdt.date().day(),
            start_qdt.time().hour(), start_qdt.time().minute(), start_qdt.time().second()
        )
        end_text = self.end_label.text().strip()
        end_dt_str = end_text if end_text not in ("—", "无法判定/无数据") else None

        record_count = None
        if self._last_wb_path and col_index and self.sheet_combo.currentText().strip():
            last_row = find_last_non_empty_row(self._last_wb_path, self.sheet_combo.currentText().strip(), col_index)
            if last_row is not None and last_row >= int(self.row_spin.value()):
                record_count = last_row - int(self.row_spin.value()) + 1

        return {
            "time_scale": "hourly",
            "pv_name": self.pv_name_edit.text().strip() or "PV",
            "file_path": self.file_edit.text().strip(),
            "sheet_name": self.sheet_combo.currentText().strip(),
            "column_letter": col_letter,
            "column_index": col_index,
            "start_row": int(self.row_spin.value()),
            "start_datetime": start_dt.strftime("%Y-%m-%d %H:%M:%S"),
            "end_datetime": end_dt_str,
            "record_count": record_count
        }

    def is_valid(self) -> bool:
        m = self.to_mapping()
        if m["file_path"]:
            return all([m["sheet_name"], m["column_letter"], m["column_index"], m["start_datetime"]])
        return True


# ------------------------------ 前端：主窗口 ------------------------------
class MappingWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("入库/光伏 数据映射配置（自动结束日期/时间）")
        self.resize(1100, 800)

        self.river_config: Dict[str, Any] = {}
        self.data_type_text: str = "daily"

        # 顶部：加载配置/保存
        self.cfg_path_edit = QLineEdit()
        self.cfg_path_edit.setPlaceholderText("请选择河流-水库配置（JSON）")
        self.cfg_path_edit.setReadOnly(True)
        self.btn_load_cfg = QPushButton("加载配置")
        self.btn_load_cfg.clicked.connect(self.on_load_config)

        self.btn_save = QPushButton("保存映射")
        self.btn_save.clicked.connect(self.on_save_mapping)
        self.btn_save.setEnabled(False)

        top_row = QHBoxLayout()
        top_row.addWidget(self.cfg_path_edit, 1)
        top_row.addWidget(self.btn_load_cfg)
        top_row.addStretch(1)
        top_row.addWidget(self.btn_save)

        # 水库区域
        self.res_scroll = QScrollArea()
        self.res_scroll.setWidgetResizable(True)
        self.res_con = QWidget()
        self.res_lay = QVBoxLayout(self.res_con)
        self.res_lay.addStretch(1)
        self.res_scroll.setWidget(self.res_con)

        res_box = QGroupBox("水库入库（按日）")
        res_box_lay = QVBoxLayout(res_box)
        res_box_lay.addWidget(QLabel("为每个水库选择：xlsx / Sheet / 列 / 起始行 / 起始日期；结束日期自动计算"))
        res_box_lay.addWidget(self.res_scroll)

        # 光伏区域（可选）
        self.enable_pv_chk = QCheckBox("启用光伏（逐小时）")
        self.enable_pv_chk.stateChanged.connect(self._toggle_pv_area)

        self.btn_add_pv = QPushButton("添加光伏序列")
        self.btn_add_pv.clicked.connect(self.on_add_pv_row)
        self.btn_add_pv.setEnabled(False)

        self.pv_scroll = QScrollArea()
        self.pv_scroll.setWidgetResizable(True)
        self.pv_con = QWidget()
        self.pv_lay = QVBoxLayout(self.pv_con)
        self.pv_lay.addStretch(1)
        self.pv_scroll.setWidget(self.pv_con)
        self.pv_scroll.setVisible(False)

        pv_box = QGroupBox("光伏数据（逐小时，可添加多个序列）")
        pv_box_lay = QVBoxLayout(pv_box)
        pv_box_lay.addWidget(QLabel("与水库相同的选择方式；起始需要“日期时间”；结束时间自动计算"))
        pv_box_lay.addWidget(self.enable_pv_chk)
        pv_box_lay.addWidget(self.btn_add_pv, alignment=Qt.AlignLeft)
        pv_box_lay.addWidget(self.pv_scroll)

        # 说明
        info = QLabel(
            "步骤：1) 加载河流-水库配置 → 2) 配置水库入库列（自动显示结束日期） → "
            "（可选）启用并添加光伏序列（自动显示结束时间） → 3) 保存映射 JSON"
        )
        info.setWordWrap(True)

        root = QVBoxLayout(self)
        root.addWidget(info)
        root.addLayout(top_row)
        root.addWidget(res_box)
        root.addWidget(pv_box)

        self._pv_count = 0  # for naming PV1, PV2, ...

    # ---------- 事件 ----------
    def on_load_config(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择河流-水库配置（JSON）", "", "JSON 文件 (*.json)")
        if not path:
            return
        try:
            cfg = load_river_config(path)
            self.river_config = cfg
            self.cfg_path_edit.setText(path)

            # data_type：daily/hourly 或中文
            dt = str(cfg.get("data_type", "daily")).strip()
            if dt in ("daily", "hourly"):
                self.data_type_text = dt
            elif dt in ("逐日", "按日"):
                self.data_type_text = "daily"
            elif dt in ("逐小时", "按小时"):
                self.data_type_text = "hourly"
            else:
                self.data_type_text = "daily"

            # 构造水库行
            self._populate_reservoir_rows(cfg)

            self.btn_save.setEnabled(True)
            QMessageBox.information(self, "成功",
                                    f"已加载配置：{os.path.basename(path)}\n"
                                    f"项目数据尺度（data_type）：{self.data_type_text}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载配置失败：\n{e}")

    def _populate_reservoir_rows(self, cfg: Dict[str, Any]):
        # 清空旧
        for i in reversed(range(self.res_lay.count())):
            w = self.res_lay.itemAt(i).widget()
            if w is not None:
                w.setParent(None)
        self.res_lay.addStretch(1)

        res_list = extract_reservoir_list(cfg)
        if not res_list:
            self.res_lay.insertWidget(self.res_lay.count() - 1, QLabel("配置中未发现任何水库。"))
            return

        for entry in res_list:
            row = ReservoirRow(entry["river_name"], entry["reservoir_name"])
            self.res_lay.insertWidget(self.res_lay.count() - 1, row)

        # 重置光伏区
        self.enable_pv_chk.setChecked(False)
        self.btn_add_pv.setEnabled(False)
        self._clear_pv_rows()

    def _toggle_pv_area(self, state: int):
        enabled = (state == Qt.Checked)
        self.pv_scroll.setVisible(enabled)
        self.btn_add_pv.setEnabled(enabled)
        if not enabled:
            self._clear_pv_rows()

    def _clear_pv_rows(self):
        for i in reversed(range(self.pv_lay.count())):
            w = self.pv_lay.itemAt(i).widget()
            if w is not None:
                w.setParent(None)
        self.pv_lay.addStretch(1)
        self._pv_count = 0

    def on_add_pv_row(self):
        self._pv_count += 1
        row = PVRow(self._pv_count)
        self.pv_lay.insertWidget(self.pv_lay.count() - 1, row)

    def on_save_mapping(self):
        # 汇总水库
        inflow_mappings: List[Dict[str, Any]] = []
        for i in range(self.res_lay.count() - 1):  # 最后一个是 stretch
            w = self.res_lay.itemAt(i).widget()
            if isinstance(w, ReservoirRow):
                if not w.is_valid():
                    QMessageBox.warning(self, "提示", f"请完善水库“{w.reservoir_name}”设置，或清空其文件路径。")
                    return
                m = w.to_mapping()
                if m["file_path"]:
                    inflow_mappings.append(m)

        # 汇总光伏（可选）
        pv_mappings: List[Dict[str, Any]] = []
        if self.enable_pv_chk.isChecked():
            for i in range(self.pv_lay.count() - 1):
                w = self.pv_lay.itemAt(i).widget()
                if isinstance(w, PVRow):
                    if not w.is_valid():
                        QMessageBox.warning(self, "提示", f"请完善光伏序列“{w.pv_name_edit.text()}”设置，或清空其文件路径。")
                        return
                    m = w.to_mapping()
                    if m["file_path"]:
                        pv_mappings.append(m)

        # 选择保存路径
        base_dir = os.path.dirname(self.cfg_path_edit.text().strip()) if self.cfg_path_edit.text().strip() else ""
        default_name = "inflow_pv_mapping.json"
        save_path, _ = QFileDialog.getSaveFileName(self, "保存映射 JSON",
                                                   os.path.join(base_dir, default_name),
                                                   "JSON 文件 (*.json)")
        if not save_path:
            return

        try:
            save_mapping(save_path, self.data_type_text, inflow_mappings, pv_mappings)
            QMessageBox.information(self, "成功", f"已保存到：\n{save_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败：\n{e}")


# ------------------------------ 入口 ------------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = MappingWindow()
    w.show()
    sys.exit(app.exec_())
